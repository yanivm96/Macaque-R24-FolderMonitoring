import pandas as pd
import json
import os
import gzip
import shutil
import re
import requests
from tabulate import tabulate
from datetime import datetime
from openpyxl import load_workbook



#SOURCE_PATH = r"/misc/work"
SOURCE_PATH = r"C:\Users\yaniv\Desktop"

METADATA_SCHEMA_PATH = os.path.join(SOURCE_PATH, 'Dropbox/Macaque R24/jsonFormats/schema.json')
METADATA_FILE_PATH = os.path.join(SOURCE_PATH, 'Dropbox/Macaque R24/subject_metadata/metadata.xlsx')
AIRR_SCHEMA_PATH = os.path.join(SOURCE_PATH, 'Dropbox/Macaque R24/jsonFormats/airr-schema.json')
GENOMIC_SCHEMA_PATH = os.path.join(SOURCE_PATH, 'Dropbox/Macaque R24/jsonFormats/genomic-schema.json')
FILE_TO_RUN_IN_PIPELINE_PATH = os.path.join(SOURCE_PATH, 'Dropbox/Macaque R24/ready_for_pipline/pipeline_files.txt') 
ALL_PIPELINE_FILES_PATH = os.path.join(SOURCE_PATH, 'Dropbox/Macaque R24/ready_for_pipline/all_pipeline_files.txt') 
PIPELINE_TABLE_PATH = os.path.join(SOURCE_PATH, 'Dropbox/Macaque R24/analysis/') 
EXCEL_FILE_PATH = os.path.join(SOURCE_PATH, 'Dropbox/Macaque R24/results/missing.xlsx')
WEBHOOK_URL = ''
NOT_FOUND = -1

with open("secrets.json", "r") as json_file: # Read the secrets from the JSON file
    details = json.load(json_file)
    WEBHOOK_URL = details["WEBHOOK_URL"]

class FolderMonitor():
    def __init__(self):
        self.new_subjects = []
        self.number_of_new_subjects = 0
        self.total_subjects = 0
        self.total_subjects_with_airr_sample = 0
        self.total_subjects_with_genomic_sample = 0
        self.total_subjects = 0
        self.total_samples_airr = 0
        self.total_samples_genomic = 0
        self.air_samples_from_past_24 = 0
        self.genomic_samples_from_past_24 = 0
        self.airr_missing_files = 0
        self.genomic_missing_files = 0
        self.subjects_missing_metadata = 0
        self.isAirr = False
        self.add_one_for_airr = False
        self.add_one_for_genomic = False
        self.past_24_sample = []
        self.reset_excel_file = True
        if not os.path.exists(ALL_PIPELINE_FILES_PATH):
        # If the file doesn't exist, create it and write some initial content
            with open(ALL_PIPELINE_FILES_PATH, "w") as new_file:
                print(ALL_PIPELINE_FILES_PATH + " created")


    # def load_counters_values(self): #loading the updated values of samples counters from the json file
    #     if os.path.exists("counters.json"):
    #         with open("counters.json", "r") as file:
    #             data = json.load(file)
    #             self.total_subjects_with_airr_sample = data.get("total_subjects_with_airr_sample", 0)
    #             self.total_subjects_with_genomic_sample = data.get("total_subjects_with_genomic_sample", 0)
    #             self.total_samples_airr = data.get("total_samples_airr", 0)
    #             self.total_samples_genomic = data.get("total_samples_genomic", 0)
    #             self.airr_missing_files = data.get("airr_missing_files", 0)
    #             self.genomic_missing_files = data.get("genomic_missing_files", 0)
    #             self.subjects_missing_metadata = data.get("subjects_missing_metadata", 0)


    # def save_counters_values(self): #saving samples counters after a run
    #     data = {
    #         "total_subjects_with_airr_sample": self.total_subjects_with_airr_sample,
    #         "total_subjects_with_genomic_sample": self.total_subjects_with_genomic_sample,
    #         "total_samples_airr": self.total_samples_airr,
    #         "total_samples_genomic": self.total_samples_genomic,
    #         "airr_missing_files": self.airr_missing_files,
    #         "genomic_missing_files": self.genomic_missing_files,
    #         "subjects_missing_metadata": self.subjects_missing_metadata
    #     }
    #     with open("counters.json", "w") as file:
    #         json.dump(data, file)
    
    # def reset_counters_values(self):
    #     self.air_samples_from_past_24 = 0
    #     self.genomic_samples_from_past_24 = 0

    def get_file_name_from_file_path(self, file_path): #getting the file name from a path
        parts = file_path.split("/")
        last_part = parts[-1]
        return last_part
    
    def get_folders_in_path(self, target_path):#getting all folders from a path
        folders = []
        items = os.listdir(target_path)
        for item in items:
            item_path = os.path.join(target_path, item)
            if os.path.isdir(item_path):
                folders.append(item_path)
        
        return folders


    def check_new_subject(self, subject_path):#the stating function of checking files and metadata
        try:
            self.add_one_for_airr = False
            self.add_one_for_genomic = False
            subject_name = self.get_file_name_from_file_path(subject_path)
            
            with open(METADATA_SCHEMA_PATH, 'r') as schema_file:
                schema = json.load(schema_file)
                required_properties = schema['required']
                data = pd.read_excel(METADATA_FILE_PATH)
                result = self.scan_subject_files(subject_path)
                row_number, missing_properties = self.scan_subject_metadata(data, required_properties, subject_name)
                self.analyze_checks_result(row_number, missing_properties, result,subject_name)
        except Exception as e:
            print("error - ", e)

        
    
    def analyze_checks_result(self, row_number, missing_properties, result,subject_name):
        row_message = metadata_message = missing_files_message = ''
        missing_airr_files, missing_genomic_files = self.split_result_to_airr_and_genomic(result)
        row_message, metadata_message = self.analyze_metadata_check_results(row_number, missing_properties,metadata_message)
        if "Not found" in row_message or metadata_message or missing_airr_files or missing_genomic_files:
            data = {
                'Subject Name': [subject_name],
                'Row': [row_message],
                'Missing metadata properties': [metadata_message],
                'Missing files - Airr': [missing_airr_files],
                'Missing files - Genomic': [missing_genomic_files]
            }

            df = pd.DataFrame(data)
            wb = load_workbook(EXCEL_FILE_PATH)
            sheet_name = "Sheet1"
            ws = wb[sheet_name]
            if self.reset_excel_file:
                self.reset_excel_file = False
                ws.delete_rows(2, ws.max_row)
            data = df.values.tolist()
            for row in data:
                ws.append(row)
            wb.save(EXCEL_FILE_PATH)
            wb.close()

    def split_result_to_airr_and_genomic(self, result):
        airr_lines = ""
        genomic_lines = ""
        for line in result:
            if 'airr' in line.lower():
                airr_lines += line + "\n"
            else:
                genomic_lines += line + "\n"

        return airr_lines, genomic_lines


    def analyze_metadata_check_results(self, row_number, missing_properties, metadata_message):
        missing_excel_data = False
        
        if row_number == NOT_FOUND:
            row_message = f"Not found\n"
            missing_excel_data = True
        else:
            row_message = f"Found in row {row_number}\n"

        if missing_properties != '':
            missing_excel_data = True
            metadata_message = f"{missing_properties}\n"   

        if missing_excel_data == True:
            self.subjects_missing_metadata += 1

        return row_message, metadata_message     

    def scan_subject_metadata(self, data, required_properties, subject_name):#checking subject metadata
        missing_properties = ''
        row_number = -1
        for index, row in data.iterrows(): # Iterate through rows and write missing properties to the file
            if row['Animal ID'] == subject_name:
                row_number = index +2
                for req in required_properties:
                    if pd.isna(data[req][index]): #checking if cell is empty
                        missing_properties +=  req + ', '

                break #finish after finding the row of the subject and checking the metadata
        
        return row_number, missing_properties
        

    def scan_subject_files(self,subject_path):
        missing = []
        schema = None 
        subject_folders = self.get_folders_in_path(subject_path)
        airr_schema = json.load(open(AIRR_SCHEMA_PATH, 'r'))
        genomic_schema = json.load(open(GENOMIC_SCHEMA_PATH, 'r'))

        for sample in subject_folders:
            sample_folders = self.get_folders_in_path(sample)
            for folder in sample_folders:
                if "airr" in self.get_file_name_from_file_path(folder):
                    if not self.add_one_for_airr:
                        self.total_subjects_with_airr_sample +=1
                        self.add_one_for_airr = True
                    self.check_if_sample_is_from_past_day(folder)
                    self.isAirr = True
                    schema = airr_schema
                else:
                    if not self.add_one_for_genomic:
                        self.total_subjects_with_genomic_sample +=1
                        self.add_one_for_genomic = True
                    self.check_if_sample_is_from_past_day(folder)
                    self.isAirr = False
                    schema = genomic_schema

                miss_res = self.check_if_folder_meets_files_required(schema, folder, subject_path, sample)

                if miss_res:
                    missing.append(miss_res)
        
        return missing
    
    def check_if_sample_is_from_past_day(self, folder):
        start_index = folder.find("/Macaque R24/sequencing")
        if start_index != -1:
            # Extract the part of the path after "Macaque R24/sequencing"
            desired_part = folder[start_index:].replace('\\', '/')
        
        if desired_part in self.past_24_sample:
            self.air_samples_from_past_24+=1
        else:
            self.total_samples_airr+=1

    def check_if_folder_meets_files_required(self, schema, folder, subject_path, sample):#checking folder files and checking if it fits to our json requerments, and returning the missings
        missing_files = ''
        required_files = schema["items"]['required']
        pipeline_files_names= []
        for required_file in required_files:
            actual_count = 0
            expected_count = schema["items"]["properties"][required_file + "_count"]["minimum"]
            for filename in os.listdir(folder):
                pattern = schema["items"]["properties"][required_file].get("pattern")
                if pattern and re.match(pattern, filename):
                    actual_count += 1
                    pipeline_files_names.append(filename)
            
            subject_name = self.get_file_name_from_file_path(subject_path)
            sample_name = self.get_file_name_from_file_path(sample)
            folder_name = self.get_file_name_from_file_path(folder)
            folder_short_name = f"{subject_name}/{sample_name}/{folder_name}"

            if actual_count < expected_count:
                missing_files +=f"{folder_short_name}: Expected {expected_count} file of type **{required_file}**, found only {actual_count}\n"

            elif expected_count != 0:
                self.manage_folder_files(folder, pipeline_files_names)

            pipeline_files_names.clear()

        if missing_files.endswith('\n'): #removing the lask \n 
            missing_files = missing_files[:-1]

        if len(missing_files) != 0:
            if self.isAirr == True:
                self.airr_missing_files+=1
            else:
                self.genomic_missing_files+=1

        return missing_files

    def manage_folder_files(self, folder, pipeline_files_names): #if the folder meets the parameters, we want to unzip the file and make them ready for pipeline
        with open(ALL_PIPELINE_FILES_PATH, "r") as all_pipeline_files:
            all_lines = all_pipeline_files.readlines()
            all_lines = [line.replace('\n', '') for line in all_lines]

        with open(FILE_TO_RUN_IN_PIPELINE_PATH, "a") as ready_for_pipeline_file:
            with open(ALL_PIPELINE_FILES_PATH, "a") as all_pipeline_files:
                for file_name in pipeline_files_names:
                    line_in_pipeline_file = os.path.join(folder, file_name)
                    if line_in_pipeline_file not in all_lines:
                        ready_for_pipeline_file.write(line_in_pipeline_file + "\n")
                        all_pipeline_files.write(line_in_pipeline_file + "\n")


    def end_of_day_summary(self):
        self.total_samples_airr +=  self.air_samples_from_past_24
        self.total_samples_genomic +=  self.genomic_samples_from_past_24
        table_data1, table_data2 = self.create_slack_table()
        table1 = tabulate(table_data1, tablefmt="fancy_grid")
        table2 = tabulate(table_data2, tablefmt="fancy_grid")
        self.send_slack_message("```\n" + table1 + "\n```")
        self.send_slack_message("```\n" + table2 + "\n```")
        


    def create_slack_table(self):
        today = datetime.now().strftime("%B %d, %Y")  # Format the date as "Month Day, Year"

        table1 = [
            [f"{today}", f"AIRR-Seq", f"Genomic"],
            ["Total Subjects", f"{self.total_subjects_with_airr_sample}", f"{self.total_subjects_with_genomic_sample}"],
            ["Total Samples", f"{self.total_samples_airr}", f"{self.total_samples_genomic}"],
            ["Samples added in past 24 hours", f"{self.air_samples_from_past_24}", f"{self.genomic_samples_from_past_24}"],
            ["Samples missing files", f"{self.airr_missing_files}", f"{self.genomic_missing_files}"]
        ]

        table2 = [
            [f"{today}", f""],
            ["Subjects missing metadata", f"{self.subjects_missing_metadata}"]
        ]

        return table1, table2
    
    def send_slack_message(self, message):
        payload = {
            "text": message
        }
        response = requests.post(WEBHOOK_URL, json=payload)


    def update_pipeline_table(self):#updetting the excel file for new subject and samples files
        columns = ["file path", "ran in pipeline"]
        data_file = os.path.join(PIPELINE_TABLE_PATH, "pipeline_table.xlsx")
        if os.path.exists(data_file):
            table_data = pd.read_excel(data_file)
        else:
            table_data = pd.DataFrame(columns=columns)
            
        with open(FILE_TO_RUN_IN_PIPELINE_PATH, "r") as pipeline_file:
            paths = pipeline_file.read().splitlines()

        path_prefixes = {}
        for object_name in paths:
            path_prefix = os.path.dirname(object_name)
            if path_prefix in path_prefixes:
                path_prefixes[path_prefix].append(object_name)
            else:
                path_prefixes[path_prefix] = [object_name]

        new_rows = []
        for path_prefix, paths in path_prefixes.items():
            new_row = {"file path": path_prefix, "ran in pipeline": False}
            new_rows.append(new_row)
        
        new_data = pd.DataFrame(new_rows)
        table_data = pd.concat([table_data, new_data], ignore_index=True)
        
        table_data.to_excel(data_file, index=False)  # Save the DataFrame to an Excel file
